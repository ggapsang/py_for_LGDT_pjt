import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Color
import random
import time
import datetime

def gen_MDM_id(load_ws) :
    print("STEP 1st : 해더값 세팅 중...")

    def making_MDM_id_conca (load_ws, mdm_id_lst, mdm_id_lst_row_no) :
        print("MDM ID 생성 중...")
        row_no = 2
        #v = 1
        for row in load_ws.iter_rows(min_row=2) :
            #print(v,"의 mdm id 생성")
            if row[mdm_upload_col].value != "X" :
                naming_codes = []
                for cell_col in load_hdr_no_dic.values() :
                    if cell_col != mdm_id_col and cell_col != mdm_upload_col and row[cell_col].value is not None :
                        naming_codes.append(str(row[cell_col].value))
                row[mdm_id_col].value = '-'.join(naming_codes) # row[key_col].value =mdmID
                mdm_id_lst.append(row[mdm_id_col].value)
                mdm_id_lst_row_no.append(row_no)
            row_no = row_no + 1
            v=v+1

    def checkingDupl(mdm_id_dict, indiv_dict) :
        print("중복 MDM ID 확인 중...")
        seen = []
        for ky, val in mdm_id_dict.items() :
            if val not in seen :
                seen.append(val)
                indiv_dict[ky] = val

        return indiv_dict

    hdr_lst = ["MDM 등록 여부", "MDM 설비 ID", "사용 유체 코드", "태그 시리얼 번호", "배관 사이즈", "배관 사양 코드", "배관 보온 코드", "배관 트레이싱 코드", "배관 자켓 코드"]
    key = "MDM 설비 ID" 
    key2 = "MDM 등록 여부"
    key3 = "태그 시리얼 번호"
    load_hdr_no_dic = {}

    # print(ldWs[1])
    for i in load_ws[1] : # 1은 해더가 있는 행의 값
        if i.value in hdr_lst and i.value not in load_hdr_no_dic.keys() :
            load_hdr_no_dic[i.value] = i.column - 1

    mdm_id_col = load_hdr_no_dic.get(key)
    mdm_upload_col = load_hdr_no_dic.get(key2)
    sr_no_col = load_hdr_no_dic.get(key3)


    print("2nd STEP : 1차 MDM ID 생성 중...")
    mdm_id_lst = []
    mdm_id_lst_row_no = []
    making_MDM_id_conca(load_ws=load_ws, mdm_id_lst=mdm_id_lst, mdm_id_lst_row_no=mdm_id_lst_row_no)
    mdm_id_dic = dict(zip(mdm_id_lst_row_no, mdm_id_lst))


    print("3rd STEP : 중복 체크 리스트 생성 중...")
    checklst = []
    indiv_dict = {}
    checkingDupl(mdm_id_dic, indiv_dict)
    checklst = [x for x in mdm_id_dic.keys() if x not in indiv_dict.keys()]
    print("mdm id 중복 행 : ", checklst)
    print("mdm id 중복 개수: ", len(checklst))


    print("4th STEP : 중복 MDM ID 변경 중...")
    row_no = 2
    while len(checklst) > 0 :
        for row_no in checklst :
            for row in load_ws.iter_rows(min_row=row_no, max_row=row_no) :
                # 중복 mdm id를 갖는 값의 serial no 변경
                sr_no_split = list(str(row[sr_no_col].value))
                new_sr_no_lst = []
                for i in sr_no_split :
                    try :
                        i = int(i)
                        i = random.randrange(0, 10)
                        new_sr_no_lst.append(str(i))
                    except :
                        new_sr_no_lst.append(str(i))
                row[sr_no_col].value = ''.join(new_sr_no_lst)
                # print(row[srNo_col].value)

                # 변경된 serial no 칼라마킹
                load_ws.cell(row_no, sr_no_col+1).fill = PatternFill(fill_type='solid', fgColor=Color('FFFF00'))

                # 변경된 serial no로 새 mdm id 생성
                naming_codes = []
                for cell_col in load_hdr_no_dic.values() :
                    if cell_col != mdm_id_col and cell_col != mdm_upload_col and row[cell_col].value is not None :
                        naming_codes.append(str(row[cell_col].value))
                row[mdm_id_col].value = '-'.join(naming_codes) # row[key_col].value =mdmID
                # print(row[mdmId_col].value)

                # mdm id 딕셔너리에 새로 반영함
                mdm_id_dic[row_no] = row[mdm_id_col].value

        # 중복 체크 루프
        checklst = []
        indiv_dict = {}
        checkingDupl(mdm_id_dic, indiv_dict)
        checklst = [x for x in mdm_id_dic.keys() if x not in indiv_dict.keys()]
        print(len(checklst))
    print("루프문 종료. 중복 태그 없음")

start = time.time()

print("1st STEP : 파일 불러오는 중..")
path = "C:/Users/dayoung.kweon/Documents/00_Python/"
load_file_name = input("파일이름 : ") + ".xlsx"
load_wb = openpyxl.load_workbook(path + load_file_name)
# loadWb.save(path + "백업파일/" + loadFilename)

for x in ["배관_BD2", "배관_BRU"] :
    load_ws = load_wb[x]
    gen_MDM_id(load_ws)

print("저장 중...")
load_wb.save(path + load_file_name)

end = time.time()
sec = end - start
result = str(datetime.timedelta(seconds=sec)).split(".")
print(result[0], "완료")
