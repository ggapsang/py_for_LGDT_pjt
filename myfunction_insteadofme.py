import openpyxl
import random
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Color
from collections import Counter

def merge_partial(hd_Ws, load_Ws, save_Ws) :

    ld_hd_row = int(input("load data의 해더 행 : "))
    sv_hd_row = int(input("save data의 해더 행 : "))
    writter = "작성자"
    lup_value = "SR No"

    ## 해더 리스트 생성
    print("해더 리스트 생성 중")
    while True : # hdrRowNo 시트에서 불러올 행
        hd_row_no = input("[LOOKUP VALUE LIST]\n    1 : [작성자, SR No, 공종변경사항, 각 공종 담당 검토사항, 대표, Tag No 수정, 카테고리 이름, 클래스 이름, 타입 이름, 제외사유, MDM 등록 여부]\n    2 : [작성자, SR No, 공종변경사항, 각 공종 담당 검토사항, 카테고리 이름, 클래스 이름, 타입 이름, 제외사유, MDM 등록 여부]\n    3 : [작성자, SR No, 공종변경사항, 각 공종 담당 검토사항, 대표, Tag No 수정, 카테고리 이름, 클래스 이름, 타입 이름, 제외사유, MDM 등록 여부, 참조 TAG NO]\n    4 : [작성자, SR No, Tag No 수정, 카테고리 이름, 클래스 이름, 타입 이름, 제외사유]\n    5 : [개별속성 입력 CCT 포함]\n    6 : [개별속성 입력 개별속성 값만]\n    7 : [작성자, SR No, 공종변경사항, 각 공종 담당 검토사항, 대표, Tag No 수정, 대표 TAG, 카테고리 이름, 클래스 이름, 타입 이름, 제외사유, MDM 등록 여부]\n 1-7 중 값 입력 : ")
        rows = int(hd_row_no)
        if rows in range(1, 8) :
            break
        else :
            print("1-8 사이에 숫자만 입력하시오")
            continue
    lup_hd_lst = []
    for rows in hd_Ws[int(hd_row_no)] : # hdr_Worksheet[]의 숫자는 원하는 값이 들어 있는 해더의 열
        lup_hd_lst.append(rows.value)

    print("lupHdr_lst : ", lup_hd_lst, "\n")

    ## 로드 데이터의 해더 딕셔너리 생성
    print("로드 데이터 딕셔너리 생성 중")
    ld_hdNo_dic = {}
    for rows in load_Ws[ld_hd_row] :
        if rows.value in lup_hd_lst and rows.value not in ld_hdNo_dic.keys() :
            ld_hdNo_dic[rows.value] = rows.column - 1
    reverse_ld_hdNo_dic = dict(map(reversed, ld_hdNo_dic.items()))

    # print(reverse_ld_hdNo_dic)

    # print('ldHdrNodict : ', ldHdrNoDict, "\n")
    # print('ldHdrNodict.values() : ',ldHdrNoDict.values(), "\n")


    ## 로드 데이터의 값 딕셔너리 생성
    print("로드 데이터의 값 딕셔너리 생성 중")
    key_loadWs = []
    value_loadWs = []
    for rows in load_Ws.iter_rows(min_row= ld_hd_row+1) :
        Dic = {}
        for j in ld_hdNo_dic.values() :
            if j == ld_hdNo_dic.get(lup_value) and rows[ld_hdNo_dic.get(writter)].value is not None :    
                key_loadWs.append(rows[ld_hdNo_dic.get(lup_value)].value)
            else :
                Dic[reverse_ld_hdNo_dic.get(j)] = rows[j].value
        value_loadWs.append(Dic)

    ld_value_dic = dict(zip(key_loadWs, value_loadWs))

    ## 세이브 데이터 해더의 딕셔너리 생성
    print("세이브 데이터 해더의 딕셔너리 생성 중")
    sv_hdNo_dic = {}
    for rows in save_Ws[int(sv_hd_row)] :
        if rows.value in lup_hd_lst and rows.value not in sv_hdNo_dic.keys() :
            sv_hdNo_dic[rows.value] = rows.column - 1
    reverse_sv_hdNo_dic = dict(map(reversed, sv_hdNo_dic.items()))

    # print("ldValueDict : ", ldValueDict, "\n")
    # print("lupHdr_lst : ", lupHdr_lst, "\n")
    # print('reverse_ldHdrNoDict : ', reverse_ldHdrNoDict ,"\n")
    # print('ldHdrNodict : ', ldHdrNoDict, "\n")
    # print('ldHdrNodict.values() : ',ldHdrNoDict.values(), "\n")
    # print("svHdrNoDict : ", svHdrNoDict, "\n")
    # print("reverse_svHdrNoDict : ", reverse_svHdrNoDcit, "\n")
    # print(len(ldValueDict))

    ## 세이브 데이터에 값 입력하기
    print("세이브 데이터에 값 입력 중")
    for rows in save_Ws.iter_rows(min_row= sv_hd_row+1) :
        lupv = rows[sv_hdNo_dic.get(lup_value)].value
        if lupv in ld_value_dic.keys() :
            for j in sv_hdNo_dic.values() :
                if j != sv_hdNo_dic.get(lup_value) and j != sv_hdNo_dic.get(writter) :
                    rows[j].value = ld_value_dic.get(lupv).get(reverse_sv_hdNo_dic.get(j))
                elif j == sv_hdNo_dic.get(writter) :
                    rows[j].value = str(rows[sv_hdNo_dic.get(writter)].value) + str(ld_value_dic.get(lupv).get(reverse_sv_hdNo_dic.get(j)))

def ins_numbering_for_colormark(ws) :
    # 약어첫자+SrNO같은 그룹 중에 1부터 순서대로 번호 삽입
    # T,V,G는 같은 값을 가지더라도 1부터 다시 번호 삽입
    # V와 TS가 같이 있고 T가 없는 경우에는 V를 우선순위로 하여 번호 삽입
   
    def get_Hdr_Dict(ws, lup_hd, hd_dic, hd_row_no) : # 찾고자 하는 해더의 칼럼 값 딕셔너리 생성
        for i in ws[hd_row_no] :
            if i.value in lup_hd and i.value not in hd_dic.keys() :
                hd_dic[i.value] = i.column
    def mark_plus_1_goDown(ws, row_no, col_markNo) : # 마크넘버 열에서 바로 위 행의 값에 더하기 1을 한다
        ws.cell(row=row_no, column=col_markNo).value = ws.cell(row=row_no-1, column=col_markNo).value + 1
    def mark_plus_1_goUp(ws, row_no, col_markNo) : # 마크넘버 열에서 바로 아래 행의 값에 더하기 1을 한다
        ws.cell(row=row_no, column=col_markNo).value = ws.cell(row=row_no+1, column=col_markNo).value + 1
    def mark_1(ws, row_no, col_markNo) : # 마크넘버 열에 1을 적는다
        ws.cell(row=row_no, column=col_markNo).value = 1
    def mark_newgr(ws, row_no, col_new_grp, col_fst_srNo) : #새그룹 열에 기존 약어첫자 sr조합을 넣는다
        ws.cell(row=row_no, column=col_new_grp).value = ws.cell(row=row_no, column=col_fst_srNo).value
    def mark_newgr_v(ws, row_no, col_new_grp, col_fst_srNo) : #새그룹 열에 기존 약어첫자 sr조합에 "_v"를 추가한다
        ws.cell(row=row_no, column=col_new_grp).value = ws.cell(row=row_no, column=col_fst_srNo).value + "_V"
    def mark_newgr_x(ws, row_no, col_new_grp) : #바로 위 행의 새그룹 열 값을 복사해서 넣는다
        ws.cell(row=row_no, column=col_new_grp).value = ws.cell(row=row_no-1, column=col_new_grp).value
    def get_check_dic(ws, check_dic, col_fst_srNo, col_group_value, col_count_group) : # check dic = "약어첫자+sr no를 키 값으로, 그룹을 value로 하는 딕셔너리 생성"
        fst_sr_lst = []
        print(ws.max_row+1)
        for row in range(hd_row_no, ws.max_row+1) :
            print("딕셔너리 생성 중 :",row)
            if row == hd_row_no :
                continue
            else :
                fst_sr = ws.cell(row=row, column=col_fst_srNo).value
                if fst_sr not in fst_sr_lst :
                    fst_sr_lst.append(fst_sr)
                    grp_value_lst = []
                    for i in range(ws.cell(row=row, column=col_count_group).value) :
                        grp_value_lst.append(ws.cell(row=row+i, column=col_group_value).value)
                        check_dic[fst_sr] = grp_value_lst
                    else :
                        continue
                else :
                    continue
    def mark_V_TS(ws, row_no, hd_dic, check_dic) : # V, TS만 있는 경우 마크하기
        fst_sr = ws.cell(row=row_no, column=col_fst_srNo).value
        col_markNo = hd_dic["마크번호"]
        countgr_dic = Counter(check_dic[fst_sr])
        print(countgr_dic)


        # T그룹이 있는 경우 그냥 두고 지나감. T가 없는 경우(TS는 반드시 하나 이상 있는 경우 이 함수가 실현되므로 굳이 그 조건은 신경 안써도 됨)
        print(check_dic[fst_sr])
        if "T" not in check_dic[fst_sr] :
            #아래로 내려가면서 V로 된 그룹에 +1씩 넘버링 하고, 위로 TS개수만큼 올라가 다음 시작 수부터 +1넘버링 하고, 나머지는 X그룹에 +1 씩 넘버링 함.
           
            #자기 자신에 1번 부여
            mark_1(ws=ws, row_no=row_no, col_markNo=col_markNo)
            mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
           
            #아래로 내려가면서 V로 된 그룹에 1부터 +1씩 넘버링
            if countgr_dic["V"] !=1 :   
                for i in range(countgr_dic["V"]) :
                    if i == 0 :
                        mark_1(ws=ws, row_no=row_no, col_markNo=col_markNo)
                        mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
                    else : 
                        mark_plus_1_goDown(ws=ws, row_no=row_no+i, col_markNo=col_markNo)
                        mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
           
            #위로 올라가면서 TS로 된 그룹에
            for i in range(1 , countgr_dic["TS"]+1) :
                #입력을 시작하는 지점(i=1)에서는 V까지 넘버링 한 값(리스트에서 V의 개수)+1로 시작
                if i == 1 :
                    ws.cell(row=row_no-i, column=col_markNo).value = countgr_dic["V"] + 1
                    mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
                else :
                    mark_plus_1_goUp(ws=ws, row_no=row_no-i, col_markNo=col_markNo)
                    mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
            #나머지 그룹(X만 남음)은 그 다음 번호부터 넘버링 함
            if "X" in countgr_dic.values() :
                for i in range(1, countgr_dic["X"]+1) :
                    if i == 1 :
                        ws.cell(row=row_no+i, column=col_markNo).value = countgr_dic["V"] + countgr_dic["TS"] + 1
                    else :
                        mark_plus_1_goDown(ws=ws, row_no=row_no+i, col_markNo=col_markNo)
        else :
            mark_1(ws=ws, row_no=row_no, col_markNo=col_markNo)
            mark_newgr_v(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)



    # 작업에 필요한 열들의 리스트, 딕셔너리
    lup_hd = ["1.약어첫자+SERIAL", "새그룹", "2.그룹", "개수", "마크번호"]
    hd_dic = {}
    hd_row_no = int(input("해더가 있는 행 번호 입력 : "))

    # 찾고자 하는 해더리스트(lup_hd)의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(ws=ws,lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)

    # 각 해더리스트의 col_no를 간결하게 표현하도록 변수 정의
    col_fst_srNo = hd_dic["1.약어첫자+SERIAL"]
    col_new_grp = hd_dic["새그룹"]
    col_group_value = hd_dic["2.그룹"]
    col_count_group = hd_dic["개수"]
    col_markNo = hd_dic["마크번호"]
   
    # check dic = "약어첫자+sr no를 키 값으로, 그룹을 value로 하는 딕셔너리 생성"
    check_dic = {}
    get_check_dic(ws=ws, check_dic=check_dic, col_fst_srNo=col_fst_srNo, col_group_value=col_group_value, col_count_group=col_count_group)

    # <넘버링 작업 시작>
    # 1) 약어첫자+SrNO같은 그룹 중에 1부터 순서대로 번호 삽입
    # 2) T,V,G는 같은 값을 가지더라도 1부터 다시 번호 삽입
    # 3) V와 TS가 같이 있고 T가 없는 경우에는 V를 우선순위로 하여 번호 삽입
    # ※ 첫 번째 행(해더)과 값의 첫 행의 경우 그냥 넘어감(엑셀시트에 수기로 미리 마크번호에 1로 써 둘 것)
    for i in range(hd_row_no+2, ws.max_row+1) : #약어첫자+srNo 안에 있는 값들만 먼저 검사 시작.
        row_no = i
        #약어첫자+srNo가 다른 경우 다른 조건과 관계없이 마크번호에 "1"삽입
        if ws.cell(row=row_no, column=col_fst_srNo).value != ws.cell(row=row_no-1, column=col_fst_srNo).value :
            ws.cell(row=row_no, column=col_markNo).value = 1
            if ws.cell(row=row_no, column=col_group_value).value == "V" :
                mark_newgr_v(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
            else :
                mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
        
        #약어첫자+srNo가 같은 경우
        else :
            # 내가 G이고 앞에도 G이면+1
            if ws.cell(row=row_no, column=col_group_value).value == "G" and ws.cell(row=row_no-1, column=col_group_value).value in ["G"]:
                mark_plus_1_goDown(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
            # 내가 T인 경우, 앞이 G이면 1, T이면 +1
            elif ws.cell(row=row_no, column=col_group_value).value == "T" and ws.cell(row=row_no-1, column=col_group_value).value in ["G"] :
                mark_1(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
            elif ws.cell(row=row_no, column=col_group_value).value == "T" and ws.cell(row=row_no-1, column=col_group_value).value in ["T"] :
                mark_plus_1_goDown(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
            # 내가 TS인 경우, 앞이 G이면 1, T나 TS이면 +1
            elif ws.cell(row=row_no, column=col_group_value).value == "TS" and ws.cell(row=row_no-1, column=col_group_value).value in ["G"] :
                mark_1(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
            elif ws.cell(row=row_no, column=col_group_value).value == "TS" and ws.cell(row=row_no-1, column=col_group_value).value in ["T", "TS"] :
                mark_plus_1_goDown(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)       
            # 내가 V인 경우, 앞이 G나 T이면 1, V이면 +1, TS면 순서바꾸기
            elif ws.cell(row=row_no, column=col_group_value).value == "V" and ws.cell(row=row_no-1, column=col_group_value).value in ["G", "T"] :
                mark_1(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr_v(ws=ws, row_no=row_no, col_new_grp=col_new_grp, col_fst_srNo=col_fst_srNo)
            elif ws.cell(row=row_no, column=col_group_value).value == "V" and ws.cell(row=row_no-1, column=col_group_value).value in ["V"] :
                mark_plus_1_goDown(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr_x(ws=ws, row_no=row_no, col_new_grp=col_new_grp)
            elif ws.cell(row=row_no, column=col_group_value).value == "V" and ws.cell(row=row_no-1, column=col_group_value).value in ["TS"] :
                mark_V_TS(ws=ws, row_no=row_no, hd_dic=hd_dic, check_dic=check_dic)
            # 내가 X인 경우, 앞이 G이면 1, T, TS, V, X이면 +1
            elif ws.cell(row=row_no, column=col_group_value).value == "X" and ws.cell(row=row_no-1, column=col_group_value).value in ["G"] :
                mark_1(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr_x(ws=ws, row_no=row_no, col_new_grp=col_new_grp)
            elif ws.cell(row=row_no, column=col_group_value).value == "X" and ws.cell(row=row_no-1, column=col_group_value).value in ["T", "TS", "V", "X"] :
                mark_plus_1_goDown(ws=ws, row_no=row_no, col_markNo=col_markNo)
                mark_newgr_x(ws=ws, row_no=row_no, col_new_grp=col_new_grp)
        print(row_no, ws.cell(row=row_no, column=col_markNo).value, ws.cell(row=row_no, column=col_fst_srNo).value)
