import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Color
import random
import time
import datetime

def get_hdr_no_dict(load_ws, hdr_lst, hdr_no_dict) :
    for i in load_ws[1] : # 1: 해더가 시작되는 열
        if i.value in hdr_lst and i.value not in hdr_no_dict.keys() :
            hdr_no_dict[i.value] = i.column - 1
    rev_hdr_no_dict = dict(map(reversed, hdr_no_dict.items()))
    return rev_hdr_no_dict

def get_pms_table(load_ws, id_hdr_no, table) :
    for i in load_ws.iter_rows(min_row= 2) :
        dic = {}
        for j in id_hdr_no :
            if j != "MAT+SIZE" :
                dic[j] = i[id_hdr_no[j]].value    
        table[i[id_hdr_no.get("MAT+SIZE")].value] = dic

start = time.time()


print("1st STEP : 파일 불러오는 중..")
path = "C:/Users/dayoung.kweon/Documents/00_Python/"
load_file_name = "대산_POE_모수정리_230425_v0.8.xlsx"
load_wb = openpyxl.load_workbook(path + load_file_name)
# load_wb.save(path + "백업파일/" + load_file_name)

pms_ws = load_wb["배관 PMS"]
pms_ws_hdr = ["MATERIAL CODE", "MAT+SIZE", "PLANT", "출처명", "출처명2", "출처명3", "타입 이름", "SIZE SCOPE(MIN)", "SIZE SCOPE(MAX)", "플랜지 래이팅", "플랜지 접촉면 타입", "배관 재질", "가스켓 재질", "배관 스케줄", "배관 두께"]
pms_ws_hdr_no = {}

for i in ["배관_POE", "배관_LLDPE"] : # pip_ws = load_wb["배관"] # 시트 이름_입력 수정 값
    pip_ws = load_wb[i]  
    pip_ws_hdr = ["MDM 등록 여부", "타입 이름", "배관 사양 코드", "사이즈2", "플랜지 래이팅", "플랜지 접촉면 타입", "배관 재질", "가스켓 재질", "배관 스케줄", "배관 두께"]
    pip_ws_value = ["타입 이름", "플랜지 래이팅", "플랜지 접촉면 타입", "배관 재질", "가스켓 재질", "배관 스케줄", "배관 두께"]
    pip_ws_hdr_no = {}

    print("2nd STEP : PMS시트 해더 리스트 생성 중..")
    get_hdr_no_dict(load_ws=pms_ws, hdr_lst=pms_ws_hdr, hdr_no_dict=pms_ws_hdr_no)
    get_hdr_no_dict(load_ws=pip_ws, hdr_lst=pip_ws_hdr, hdr_no_dict=pip_ws_hdr_no)
    print(pms_ws_hdr_no)
    print(pip_ws_hdr_no)

    pms_table = {}
    get_pms_table(load_ws=pms_ws, id_hdr_no=pms_ws_hdr_no, table=pms_table)
    # print(pms_table)

    for i in pip_ws.iter_rows(min_row=2) :
        find_count = 0
        not_find_count = 0
        if i[pip_ws_hdr_no["MDM 등록 여부"]].value == "△" :
            for j in pms_table :
                try :
                    if i[pip_ws_hdr_no["배관 사양 코드"]].value == pms_table[j].get("MATERIAL CODE") and pms_table[j].get("SIZE SCOPE(MIN)") <= i[pip_ws_hdr_no["사이즈2"]].value and i[pip_ws_hdr_no["사이즈2"]].value <= pms_table[j].get("SIZE SCOPE(MAX)") :
                        find_count = find_count+1
                        for k in pip_ws_value :
                            i[pip_ws_hdr_no[k]].value = pms_table[j].get(k)
                except :
                    not_find_count = not_find_count + 1
                    continue
        else : 
            continue            
        # print("찾은속성값 : ", find_count," 못찾은 속성값 : ", not_find_count)

    # print(pms_table)

print("저장 중...")
load_wb.save(path + load_file_name)

end = time.time()
sec = end - start
result = str(datetime.timedelta(seconds=sec)).split(".")
print(result[0], "완료")
