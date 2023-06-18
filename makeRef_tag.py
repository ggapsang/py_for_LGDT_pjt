import openpyxl
import time
import datetime

def genrefTag(loadWs) :
    loadWsHdr = ['Tag No 수정', '대표 TAG', 'MDM 등록 여부', '참조 TAG NO']

    loadFileHdr = []
    for cell in loadWs[11] : # 해더가 있는 행의 번호를 적는다
        loadFileHdr.append(cell.value)
    print(loadFileHdr)
    
    hdrTpNo = []
    for i in loadWsHdr :
        if i in loadFileHdr :
            hdrTpNo.append(loadFileHdr.index(i))
    if len(hdrTpNo) != 4 :
        print("못 찾은 해더가 있음")
        print(hdrTpNo)
        exit()
    
    # for i in range(0, len(hdrTpNo)) :
    #     if hdrTpNo[0] > hdrTpNo[i] or hdrTpNo[3] < hdrTpNo[i] :
    #         print("올바른 순서로 정렬-tag no 수정과 ref tag no 행은 시작과 끝에 있어야 함")
    #         print(hdrTpNo)
    #         exit()
    #     else :
    #         continue

    print(hdrTpNo)

    tagClnmNo = hdrTpNo[0] + 1
    repClnmNo = hdrTpNo[1] + 1
    mdmUploadClnmNo = hdrTpNo[2] + 1
    refClnmNo = hdrTpNo[3] + 1

    itrTagTpNo = 0
    itrRepTpNo = repClnmNo - tagClnmNo #1 # print(itrRepTpNo)
    itrRefTpNo = refClnmNo - tagClnmNo #11 # print(itrRefTpNo)
    itrmdmUpload = mdmUploadClnmNo -tagClnmNo

    tagKey = []
    repValue = []

    # Tag의 튜플 No=0, 대표태그의 튜플 No=2, 레퍼런스 태그의 튜플 No=10
    for i in loadWs.iter_rows(min_row=12, min_col=tagClnmNo, max_col=refClnmNo) :
        if i[itrRepTpNo].value is not None and i[itrmdmUpload].value != 'X' :
            tagKey.append(i[itrTagTpNo].value)
            repValue.append(i[itrRepTpNo].value)

    refTag_dict = dict(zip(tagKey, repValue))

    # Tag의 튜플 No=0, 대표태그의 튜플 No=2, 레퍼런스 태그의 튜플 No=10
    for i in loadWs.iter_rows(min_row=12, min_col=tagClnmNo, max_col=refClnmNo) :
        if i[itrTagTpNo].value in refTag_dict.keys() and i[itrTagTpNo].value == i[itrRepTpNo].value :
            indivRefLst = []
            for j in refTag_dict.keys() :
                if j != i[itrRepTpNo].value and refTag_dict[j] == i[itrRepTpNo].value :
                    indivRefLst.append(j)
                i[itrRefTpNo].value = ','.join(indivRefLst)
        else :
            continue
    print('Reference Tag 생성 완료')

start = time.time()

print("파일 불러오는 중")

path = "C:/Users/dayoung.kweon/Documents/00_Python/"
loadFilename = "작업용_여수(용성)_아크릴레이트1팀_통합설비리스트_20230411_v1.5_relation 반영완료.xlsx"

loadWb = openpyxl.load_workbook(path + loadFilename)
loadWb.save(path + loadFilename)

print("본 작업 진행 중")
for i in ["배관 외_3AA"] :
    loadWs = loadWb[i]
    genrefTag(loadWs=loadWs)


print("저장 중")
loadWb.save(path + loadFilename)
end = time.time()
sec = end - start
result = str(datetime.timedelta(seconds=sec)).split(".")
print(result[0], "완료")
