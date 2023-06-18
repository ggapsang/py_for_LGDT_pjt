import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Color
import time
from collections import Counter
import datetime

# 찾고자 하는 해더의 튜플 값 딕셔너리 생성
def get_Hdr_Dict(Ws, lup_hd, hd_dic, hd_row_no) :
    for i in Ws[int(hd_row_no)] :
        if i.value in lup_hd and i.value not in hd_dic.keys() :
            hd_dic[i.value] = i.column - 1

# Validation 항목 열에 원하는 멘트 추가하기
def fill_Validation_co(rows, hd_dic, contents) :
    if rows[hd_dic["VALIDATION"]].value is None :
       rows[hd_dic["VALIDATION"]].value = contents
    else :
       rows[hd_dic["VALIDATION"]].value = str(rows[hd_dic["VALIDATION"]].value) + "/" + contents

# 1_CHECK : MDM 등록여부 항목에 공백이 있는지 확인
def isblank_MDMupLoad(Ws) :
    lup_hd = ["VALIDATION", "SR No", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        # Check 1의 검증
        if rows[hd_dic["MDM 등록 여부"]].value is None :
            # Comment항목 열에 원하는 멘트 추가하기
            fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="1_MDM 등록 여부 공백")
        else :
            continue

# 2_CHECK : "SR NO"와 "대표" 값 비교 결과 다르면, MDM 등록 여부는 REF 또는 X여야 함.
def isfalse_MDMupLoad_1(Ws) :
    lup_hd = ["VALIDATION", "SR No", "대표", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        # CHECK 2의 검증
        if rows[hd_dic["SR No"]].value != rows[hd_dic["대표"]].value and rows[hd_dic["MDM 등록 여부"]].value not in ["x", "X", "REF"] :
            # Comment항목 열에 원하는 멘트 추가하기
            fill_Validation_co(rows=rows, hd_dic=hd_dic, contents= "2_SR NO !=대표, MDM등록여부 O")
        else :
            continue

# 3_CHECK : 계기가 아닌 다른 공종의 경우 "SR No"와 "대표" 값 비교 결과 다르면, "MDM등록여부"는 X이고, 항목에는 "중복"이라 되어 있어야 함
def isfalse_MDMupLoad_2(Ws) :
    lup_hd = ["VALIDATION", "SR No", "대표", "확정공종", "제외사유", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        # CHECK 3의 검증 - SR No와 대표sr이 다르고 계기 태그가 아닌데 제외사유 항목에 중복이라 안 되어 있는 경우
        if rows[hd_dic["SR No"]].value != rows[hd_dic["대표"]].value and rows[hd_dic["확정공종"]].value != "계기" :
                try :
                    if "중복" not in rows[hd_dic["제외사유"]].value :
                        # Comment항목 열에 원하는 멘트 추가하기
                        fill_Validation_co(rows=rows, hd_dic=hd_dic, contents= "3_제외사유 확인")
                    else :
                        continue
                except :
                    fill_Validation_co(rows=rows, hd_dic=hd_dic, contents= "3_제외사유 확인 중 오류")
        else :
            continue

# 4_CHECK : 대표 태그는 MDM등록 여부에 O로 표시되어 있어야 함(아닌 경우 제외사유 확인)
def isRepTag_MDMupLoad(Ws) :
    lup_hd = ["VALIDATION", "SR No", "대표", "제외사유", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    # SR No 딕셔너리 생성(key : SR No, value : mdm 등록 여부)
    srNo_mdmUp = {}
    for rows in Ws.iter_rows(min_row= int(hd_row_no)+1) :
        srNo_mdmUp[rows[hd_dic["SR No"]].value] = rows[hd_dic["MDM 등록 여부"]].value
        
    for rows in Ws.iter_rows(min_row= int(hd_row_no) + 1) :
        # 4_CHECK 검증 - 중복제외된 태그의 대표 태그는 MDM 등록 여부에 O로 되어 있어야 함
        # "SR != 대표"인데, 대표 태그가 MDM 등록 대상이 아니라고 되어 있는 경우
        try :
            if rows[hd_dic["SR No"]].value != rows[hd_dic["대표"]].value and srNo_mdmUp[rows[hd_dic["대표"]].value] not in ["o", "O", "△"] :
                # MDM 등록 대상이 아닌 것으로 표시된 대표 태그의 제외사유 확인
                j = 0
                for because_del in ["ALARM", "DCS", "SOFTWARE", "STATUS", "SWITCH", "DUMMY", "비관리대상", "불완전", "태그아님", "INTERLOCK", "MANUAL VALVE", "추출오류", "THERMOWELL"] :
                    if because_del in rows[hd_dic["제외사유"]].value :
                        j = j + 1
                    else :
                        continue
                if j > 0 :
                    # Comment항목 열에 원하는 멘트 추가하기
                    fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="4.1_대표 태그의 MDM 등록 여부 확인")
                else : 
                    # Comment항목 열에 원하는 멘트 추가하기
                    fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="4.2_제외사유 수정(중복 외 다른 것으로)")
        except :
            fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="4.3_해당 SR NO 찾을 수 없음")

# 5_CHECK : 제외 태그인데 제외 사유가 공란인 경우
def isblank_MDMupXreason(Ws) :
    lup_hd = ["VALIDATION", "SR No", "대표", "클래스 이름", "제외사유", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        # CHECK 5의 검증
        if rows[hd_dic["클래스 이름"]].value == "제외" and rows[hd_dic["제외사유"]].value is None :
            # Comment항목 열에 원하는 멘트 추가하기
            fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="5_제외사유 항목 공란")
        else :
            continue

# 6_CHECK : DUPL 체크1
def isdupl_Tag_1(Ws) :
    lup_hd = ["VALIDATION", "Tag No", "Tag No 수정", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    # mdm 등록 여부가 O인 태그 NO 수정 리스트 생성
    rev_tagNo = []
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        if rows[hd_dic["MDM 등록 여부"]].value in ["o", "O", "△", "△(모름)", "O(모름)", "o"] :
            rev_tagNo.append(rows[hd_dic["Tag No 수정"]].value)
    
    # 태그 NO 수정 리스트의 중복값 확인
    count_rev_tagNo = dict(Counter(rev_tagNo))
    duplLst_tagNo = []
    for tag in count_rev_tagNo :
        if count_rev_tagNo[tag] > 1 :
            duplLst_tagNo.append(tag)

    # 중복 태그들에 대하여 체크 표시
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        if rows[hd_dic["Tag No 수정"]].value in duplLst_tagNo :
            # Comment항목 열에 원하는 멘트 추가하기
            fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="6_중복 체크")
        else :
            continue

# 7_CHECK : DUPL 체크2_표기오류. []가 필요 없이 들어가 있는 경우
def isdupl_Tag_2(Ws) :
    lup_hd = ["VALIDATION", "Tag No", "Tag No 수정", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    # Tag No 수정 중 mdm 등록 여부에 o인 리스트 생성
    rev_tagNo = []
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        if rows[hd_dic["MDM 등록 여부"]].value in ["o", "O", "△"] :
            rev_tagNo.append(rows[hd_dic["Tag No 수정"]].value)
    print("rev_tagNo : ", rev_tagNo)

    # tag no 수정 항목에서 듀플 표시 [?]이 있는 리스트를 추출
    duplLst_tagNo = []
    for tag in rev_tagNo :
        if "[" in tag :
            duplLst_tagNo.append(tag)
    print("duplLst_tagNo : ",duplLst_tagNo)

    # [?]표시를 지운 태그 No 리스트를 생성
    del_duplSign_tagNo = []
    for tag in duplLst_tagNo :
        del_dupl_tag = tag[0 : len(tag)-3]
        del_duplSign_tagNo.append(del_dupl_tag)
    print("del_duplSign_tagNo :", del_duplSign_tagNo)
    
    # 듀플 표시를 지운 값들의 개수를 확인하여 딕셔너리에 담음
    count_del_duplSign = dict(Counter(del_duplSign_tagNo))
    print("count_del_duplSign : ", count_del_duplSign)
    falut_list = []
    # 7_check 검증
    for tag in del_duplSign_tagNo :
        # dupl 사인을 지운 tag no가 한개밖에 없는 경우 fault list에 담음: [?]가 없어야 하는 태그
        if count_del_duplSign[tag] == 1 :
           falut_list.append(tag) 
        else :
            # dupl tag no 리스트에서, 듀플 사인을 지운 태그와 일치할 때마다 끝에 있는 듀플 숫자를 땐다. 이 숫자들의 최대값은 듀플 사인을 지운 태그 개수와 같아야 하고, 이 숫자들의 합은 최대값의 1부터의 등차수열의 합 (n+1)n/2와 일치해야 한다.
            #dup_no리스트를 만들어 [?]를 뗐을 때, 일치하는 태그 리스트의 듀플 숫자값만 가져옴 
            dupl_no = []
            for dupl_tag in duplLst_tagNo :
                if tag == dupl_tag[0 : len(dupl_tag)-3] :
                    dupl_no.append(int(dupl_tag[len(dupl_tag)-2: len(dupl_tag)-1]))
                else :
                    continue
        # 이 리스트에서 리스트 원소들의 최대값은 듀플 사인을 지운 태그 개수와 같아야 하고, 이 숫자들의 합은 최대값의 1부터의 등차수열의 합 (n+1)n/2와 일치해야 한다.
            if max(dupl_no) == len(dupl_no) and (sum(dupl_no) == (max(dupl_no))*(max(dupl_no)+1)/2) :
                continue
            else :
                falut_list.append(tag)

    print("CHECK LIST : ", falut_list)
    # 중복 표기 오류 태그들에 대하여 체크 표시
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        checktag = rows[hd_dic["Tag No"]].value
        if checktag in falut_list :
            # Comment항목 열에 원하는 멘트 추가하기
            fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="7_dupl 표기 체크")
        else :
            continue

# 8_CHECK : REF 태그에 듀플 표시가 되어 있는 경우
def isRefTag_has_duplSign(Ws) :
    lup_hd = ["VALIDATION", "Tag No 수정", "MDM 등록 여부"]
    hd_dic = {}
    hd_row_no = input("해더가 있는 행 번호 입력 : ")
    
    # 찾고자 하는 해더의 튜플 값 딕셔너리 생성
    get_Hdr_Dict(Ws=Ws, lup_hd=lup_hd, hd_dic=hd_dic, hd_row_no=hd_row_no)
    
    for rows in Ws.iter_rows(min_row=int(hd_row_no)+1) :
        # CHECK 8의 검증
        if rows[hd_dic["MDM 등록 여부"]].value == "REF" and "[" in rows[hd_dic["Tag No 수정"]].value :
            # Comment항목 열에 원하는 멘트 추가하기
            fill_Validation_co(rows=rows, hd_dic=hd_dic, contents="8_REF에 듀플 표시됨")
        else :
            continue

def do_all_check(ws) :
    isblank_MDMupLoad(Ws=ws)
    isfalse_MDMupLoad_1(Ws=ws)
    isfalse_MDMupLoad_2(Ws=ws)
    isRepTag_MDMupLoad(Ws=ws)
    isblank_MDMupXreason(Ws=ws)
    isdupl_Tag_1(Ws=ws)
    isRefTag_has_duplSign(Ws=ws)



